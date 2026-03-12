#!/usr/bin/env python3
"""USALI workbook ingestion pipeline.

Usage:
    python usali_pipeline.py ingest --input "USALI Statement of Accounts - NOV P&L v2.4.xlsx" --db usali.db
"""

from __future__ import annotations

import argparse
import datetime as dt
import csv
import re
import sqlite3
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

SUMMARY_SHEETS = {
    "For Operators",
    "For Owners",
    "Rooms",
    "F&B",
    "F&B R1",
    "F&B R2",
    "F&B R3",
    "F&B R4",
    "F&B R5",
    "F&B R6",
    "F&B R7",
    "F&B B1",
    "F&B B2",
    "F&B B3",
    "F&B B4",
    "F&B In-Room",
    "F&B Banquet",
    "F&B Mini Bar",
    "Other Operated Dep",
    "Golf",
    "Spa",
    "Parking",
    "A&G",
    "Guest Laundry",
    "Minor Operated Dep",
    "Misc Income",
    "Infos and Telecome",
    "S&M",
    "POAM",
    "Utilities",
    "Management Fees",
    "Non OP",
    "House Laundry",
    "Staff Canteen",
    "Payroll-Related",
}

ACCOUNT_RE = re.compile(r"\d{6}")
COST_CENTER_RE = re.compile(r"\d{4}")


def to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_div(num, den):
    if num is None or den in (None, 0):
        return None
    return num / den


def init_db(conn: sqlite3.Connection):
    conn.executescript(
        """
        PRAGMA foreign_keys=ON;

        CREATE TABLE IF NOT EXISTS imports (
          id INTEGER PRIMARY KEY,
          file_name TEXT NOT NULL,
          period TEXT,
          uploaded_month TEXT,
          financial_year_start TEXT,
          bu_code TEXT,
          budget_ledger TEXT,
          forecast_version TEXT,
          imported_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS statement_lines (
          id INTEGER PRIMARY KEY,
          import_id INTEGER NOT NULL REFERENCES imports(id) ON DELETE CASCADE,
          sheet_name TEXT NOT NULL,
          section TEXT,
          row_no INTEGER NOT NULL,
          line_label TEXT,
          account_expression TEXT,
          department_expression TEXT,
          revenue_expression TEXT,
          actual REAL,
          actual_pct REAL,
          forecast REAL,
          forecast_pct REAL,
          budget REAL,
          budget_pct REAL,
          prior_year REAL,
          prior_year_pct REAL,
          variance_budget REAL,
          variance_budget_pct REAL,
          variance_prior_year REAL,
          variance_prior_year_pct REAL
        );

        CREATE TABLE IF NOT EXISTS account_facts (
          id INTEGER PRIMARY KEY,
          import_id INTEGER NOT NULL REFERENCES imports(id) ON DELETE CASCADE,
          block_index INTEGER NOT NULL,
          scenario_period TEXT,
          scenario_ledger TEXT,
          account_code TEXT,
          cost_center_code TEXT,
          market_segment_code TEXT,
          clubbing_code TEXT,
          amount REAL
        );

        CREATE TABLE IF NOT EXISTS chart_of_accounts (
          account_code TEXT PRIMARY KEY,
          first_seen_import INTEGER NOT NULL REFERENCES imports(id),
          last_seen_import INTEGER NOT NULL REFERENCES imports(id)
        );

        CREATE TABLE IF NOT EXISTS cost_centers (
          cost_center_code TEXT PRIMARY KEY,
          first_seen_import INTEGER NOT NULL REFERENCES imports(id),
          last_seen_import INTEGER NOT NULL REFERENCES imports(id)
        );

        CREATE TABLE IF NOT EXISTS account_cost_center_map (
          account_code TEXT NOT NULL,
          cost_center_code TEXT NOT NULL,
          first_seen_import INTEGER NOT NULL REFERENCES imports(id),
          last_seen_import INTEGER NOT NULL REFERENCES imports(id),
          PRIMARY KEY (account_code, cost_center_code)
        );

        CREATE VIEW IF NOT EXISTS v_statement_trend AS
        SELECT
          i.period,
          s.sheet_name,
          s.line_label,
          SUM(s.actual) AS actual,
          SUM(s.budget) AS budget,
          SUM(s.prior_year) AS prior_year,
          SUM(s.variance_budget) AS variance_budget,
          SUM(s.variance_prior_year) AS variance_prior_year
        FROM statement_lines s
        JOIN imports i ON i.id = s.import_id
        WHERE s.line_label IS NOT NULL
        GROUP BY i.period, s.sheet_name, s.line_label;
        """
    )

    existing_import_cols = {row[1] for row in conn.execute("PRAGMA table_info(imports)")}
    if "uploaded_month" not in existing_import_cols:
        conn.execute("ALTER TABLE imports ADD COLUMN uploaded_month TEXT")

    conn.commit()


def insert_import(conn: sqlite3.Connection, wb, file_path: Path, upload_month: str | None = None) -> int:
    start = wb["Start"]
    period = start[5][1].value if start.max_row >= 5 else None
    financial_year_start = start[9][1].value if start.max_row >= 9 else None
    bu_code = start[4][1].value if start.max_row >= 4 else None
    budget_ledger = start[10][1].value if start.max_row >= 10 else None
    forecast_version = start[8][1].value if start.max_row >= 8 else None

    cur = conn.execute(
        """
        INSERT INTO imports(file_name, period, uploaded_month, financial_year_start, bu_code, budget_ledger, forecast_version, imported_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            file_path.name,
            str(period) if period is not None else None,
            upload_month,
            str(financial_year_start) if financial_year_start is not None else None,
            str(bu_code) if bu_code is not None else None,
            str(budget_ledger) if budget_ledger is not None else None,
            str(forecast_version) if forecast_version is not None else None,
            dt.datetime.utcnow().isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    return cur.lastrowid


def parse_statement_sheet(conn, import_id: int, ws):
    section = None
    for r in range(11, ws.max_row + 1):
        account_expr = ws.cell(r, 1).value
        dept_expr = ws.cell(r, 2).value
        rev_expr = ws.cell(r, 3).value
        cdf_actual = to_float(ws.cell(r, 4).value)
        pct_actual = to_float(ws.cell(r, 5).value)
        cdf_forecast = to_float(ws.cell(r, 6).value)
        pct_forecast = to_float(ws.cell(r, 7).value)
        cdf_budget = to_float(ws.cell(r, 8).value)
        pct_budget = to_float(ws.cell(r, 9).value)
        cdf_prior = to_float(ws.cell(r, 10).value)
        pct_prior = to_float(ws.cell(r, 11).value)
        label = ws.cell(r, 12).value

        if isinstance(ws.cell(r, 4).value, str) and not label:
            section = ws.cell(r, 4).value
            continue

        if label is None and not any(
            v is not None for v in [cdf_actual, cdf_forecast, cdf_budget, cdf_prior]
        ):
            continue

        variance_budget = None if cdf_actual is None or cdf_budget is None else cdf_actual - cdf_budget
        variance_prior = None if cdf_actual is None or cdf_prior is None else cdf_actual - cdf_prior

        conn.execute(
            """
            INSERT INTO statement_lines(
              import_id, sheet_name, section, row_no, line_label,
              account_expression, department_expression, revenue_expression,
              actual, actual_pct, forecast, forecast_pct,
              budget, budget_pct, prior_year, prior_year_pct,
              variance_budget, variance_budget_pct, variance_prior_year, variance_prior_year_pct
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                import_id,
                ws.title,
                section,
                r,
                str(label) if label is not None else None,
                str(account_expr) if account_expr is not None else None,
                str(dept_expr) if dept_expr is not None else None,
                str(rev_expr) if rev_expr is not None else None,
                cdf_actual,
                pct_actual,
                cdf_forecast,
                pct_forecast,
                cdf_budget,
                pct_budget,
                cdf_prior,
                pct_prior,
                variance_budget,
                safe_div(variance_budget, cdf_budget),
                variance_prior,
                safe_div(variance_prior, cdf_prior),
            ),
        )


def master_extract_blocks(ws) -> Iterable[tuple[int, int]]:
    # block starts one column before 'Analysis Code' and has 4 useful columns: acc, cc, seg, club, amount
    for c in range(1, ws.max_column + 1):
        if ws.cell(8, c).value == "Analysis Code" and ws.cell(8, c + 3).value == "Base Amount":
            yield c - 1, c


def parse_master_extract(conn, import_id: int, ws):
    for block_index, (start_col, hdr_col) in enumerate(master_extract_blocks(ws), start=1):
        scenario_period = ws.cell(1, start_col).value
        scenario_ledger = ws.cell(2, start_col).value
        for r in range(9, ws.max_row + 1):
            account = ws.cell(r, start_col).value
            amount = to_float(ws.cell(r, start_col + 4).value)
            if account is None and amount is None:
                continue
            cc = ws.cell(r, start_col + 1).value
            seg = ws.cell(r, start_col + 2).value
            club = ws.cell(r, start_col + 3).value

            account_s = str(account) if account is not None else None
            cc_s = str(cc) if cc is not None else None
            seg_s = str(seg) if seg is not None else None
            club_s = str(club) if club is not None else None

            conn.execute(
                """
                INSERT INTO account_facts(
                  import_id, block_index, scenario_period, scenario_ledger,
                  account_code, cost_center_code, market_segment_code, clubbing_code, amount
                ) VALUES (?,?,?,?,?,?,?,?,?)
                """,
                (
                    import_id,
                    block_index,
                    str(scenario_period) if scenario_period is not None else None,
                    str(scenario_ledger) if scenario_ledger is not None else None,
                    account_s,
                    cc_s,
                    seg_s,
                    club_s,
                    amount,
                ),
            )


def update_coa_dimensions(conn, import_id: int):
    accounts = conn.execute(
        "SELECT DISTINCT account_code FROM account_facts WHERE import_id=? AND account_code IS NOT NULL",
        (import_id,),
    ).fetchall()
    for (acc,) in accounts:
        if not ACCOUNT_RE.search(acc):
            continue
        conn.execute(
            """
            INSERT INTO chart_of_accounts(account_code, first_seen_import, last_seen_import)
            VALUES (?, ?, ?)
            ON CONFLICT(account_code) DO UPDATE SET last_seen_import=excluded.last_seen_import
            """,
            (acc, import_id, import_id),
        )

    cost_centers = conn.execute(
        "SELECT DISTINCT cost_center_code FROM account_facts WHERE import_id=? AND cost_center_code IS NOT NULL",
        (import_id,),
    ).fetchall()
    for (cc,) in cost_centers:
        if not COST_CENTER_RE.search(cc):
            continue
        conn.execute(
            """
            INSERT INTO cost_centers(cost_center_code, first_seen_import, last_seen_import)
            VALUES (?, ?, ?)
            ON CONFLICT(cost_center_code) DO UPDATE SET last_seen_import=excluded.last_seen_import
            """,
            (cc, import_id, import_id),
        )

    mappings = conn.execute(
        """
        SELECT DISTINCT account_code, cost_center_code
        FROM account_facts
        WHERE import_id=? AND account_code IS NOT NULL AND cost_center_code IS NOT NULL
        """,
        (import_id,),
    ).fetchall()
    for acc, cc in mappings:
        if not ACCOUNT_RE.search(acc) or not COST_CENTER_RE.search(cc):
            continue
        conn.execute(
            """
            INSERT INTO account_cost_center_map(account_code, cost_center_code, first_seen_import, last_seen_import)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(account_code, cost_center_code)
            DO UPDATE SET last_seen_import=excluded.last_seen_import
            """,
            (acc, cc, import_id, import_id),
        )


def ingest(input_file: Path, db_path: Path, upload_month: str | None = None):
    wb = load_workbook(input_file, data_only=True)
    conn = sqlite3.connect(db_path)
    init_db(conn)
    import_id = insert_import(conn, wb, input_file, upload_month=upload_month)

    for sheet in wb.sheetnames:
        if sheet in SUMMARY_SHEETS:
            parse_statement_sheet(conn, import_id, wb[sheet])

    parse_master_extract(conn, import_id, wb["Master_Extract"])
    update_coa_dimensions(conn, import_id)

    conn.commit()
    conn.close()
    print(f"Ingested import_id={import_id} into {db_path}")
    return import_id


def latest_import_id(conn: sqlite3.Connection) -> int | None:
    row = conn.execute("SELECT MAX(id) FROM imports").fetchone()
    return row[0] if row and row[0] is not None else None


def export_query_to_csv(conn: sqlite3.Connection, query: str, params: tuple, output_path: Path):
    rows = conn.execute(query, params).fetchall()
    headers = [d[0] for d in conn.execute(query, params).description]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return len(rows)


def export_verification_data(db_path: Path, output_dir: Path, import_id: int | None):
    conn = sqlite3.connect(db_path)
    use_import_id = import_id if import_id is not None else latest_import_id(conn)
    if use_import_id is None:
        raise ValueError("No imports found in the database. Run ingest first.")

    files = {
        "chart_of_accounts.csv": (
            """
            SELECT account_code, first_seen_import, last_seen_import
            FROM chart_of_accounts
            ORDER BY account_code
            """,
            (),
        ),
        "cost_centers.csv": (
            """
            SELECT cost_center_code, first_seen_import, last_seen_import
            FROM cost_centers
            ORDER BY cost_center_code
            """,
            (),
        ),
        "account_cost_center_map.csv": (
            """
            SELECT account_code, cost_center_code, first_seen_import, last_seen_import
            FROM account_cost_center_map
            ORDER BY account_code, cost_center_code
            """,
            (),
        ),
        "account_facts_latest_import.csv": (
            """
            SELECT import_id, block_index, scenario_period, scenario_ledger,
                   account_code, cost_center_code, market_segment_code, clubbing_code, amount
            FROM account_facts
            WHERE import_id = ?
            ORDER BY block_index, account_code, cost_center_code
            """,
            (use_import_id,),
        ),
        "statement_lines_latest_import.csv": (
            """
            SELECT import_id, sheet_name, section, row_no, line_label,
                   account_expression, department_expression, revenue_expression,
                   actual, budget, prior_year,
                   variance_budget, variance_prior_year
            FROM statement_lines
            WHERE import_id = ?
            ORDER BY sheet_name, row_no
            """,
            (use_import_id,),
        ),
    }

    for file_name, (query, params) in files.items():
        row_count = export_query_to_csv(conn, query, params, output_dir / file_name)
        print(f"Wrote {row_count} rows -> {output_dir / file_name}")

    conn.close()


def verify_import(db_path: Path, import_id: int | None):
    conn = sqlite3.connect(db_path)
    use_import_id = import_id if import_id is not None else latest_import_id(conn)
    if use_import_id is None:
        raise ValueError("No imports found in the database. Run ingest first.")

    print(f"Verification for import_id={use_import_id}")
    checks = {
        "statement_lines": "SELECT COUNT(*) FROM statement_lines WHERE import_id=?",
        "account_facts": "SELECT COUNT(*) FROM account_facts WHERE import_id=?",
        "distinct_accounts": "SELECT COUNT(DISTINCT account_code) FROM account_facts WHERE import_id=? AND account_code IS NOT NULL",
        "distinct_cost_centers": "SELECT COUNT(DISTINCT cost_center_code) FROM account_facts WHERE import_id=? AND cost_center_code IS NOT NULL",
        "mapped_account_cost_centers": "SELECT COUNT(DISTINCT account_code || '|' || cost_center_code) FROM account_facts WHERE import_id=? AND account_code IS NOT NULL AND cost_center_code IS NOT NULL",
    }
    for name, query in checks.items():
        value = conn.execute(query, (use_import_id,)).fetchone()[0]
        print(f"  {name}: {value}")

    print("\nTop 20 chart_of_accounts:")
    for row in conn.execute("SELECT account_code FROM chart_of_accounts ORDER BY account_code LIMIT 20"):
        print(" ", row[0])

    print("\nTop 20 cost_centers:")
    for row in conn.execute("SELECT cost_center_code FROM cost_centers ORDER BY cost_center_code LIMIT 20"):
        print(" ", row[0])

    print("\nSample account facts (latest import):")
    for row in conn.execute(
        """
        SELECT block_index, scenario_period, scenario_ledger, account_code, cost_center_code, amount
        FROM account_facts
        WHERE import_id=?
        ORDER BY block_index, account_code, cost_center_code
        LIMIT 20
        """,
        (use_import_id,),
    ):
        print(" ", row)

    conn.close()


def main():
    parser = argparse.ArgumentParser(description="USALI P&L extraction and loading pipeline")
    sub = parser.add_subparsers(dest="command", required=True)

    ingest_cmd = sub.add_parser("ingest", help="Ingest a workbook into SQLite")
    ingest_cmd.add_argument("--input", required=True, type=Path)
    ingest_cmd.add_argument("--db", default=Path("usali.db"), type=Path)
    ingest_cmd.add_argument("--upload-month", default=None, help="Upload month label, e.g. 2025-11")

    verify_cmd = sub.add_parser("verify", help="Verify extracted data and print summary checks")
    verify_cmd.add_argument("--db", default=Path("usali.db"), type=Path)
    verify_cmd.add_argument("--import-id", type=int, default=None)

    export_cmd = sub.add_parser("export-verification", help="Export extracted data to CSV for manual verification")
    export_cmd.add_argument("--db", default=Path("usali.db"), type=Path)
    export_cmd.add_argument("--outdir", default=Path("verification_exports"), type=Path)
    export_cmd.add_argument("--import-id", type=int, default=None)

    args = parser.parse_args()
    if args.command == "ingest":
        ingest(args.input, args.db, upload_month=args.upload_month)
    elif args.command == "verify":
        verify_import(args.db, args.import_id)
    elif args.command == "export-verification":
        export_verification_data(args.db, args.outdir, args.import_id)


if __name__ == "__main__":
    main()
